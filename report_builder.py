"""Renders the HTML report and produces JSON/CSV/XLSX exports from audit data."""
from __future__ import annotations

import csv
import io
import json
from pathlib import Path

from jinja2 import Environment, FileSystemLoader, select_autoescape

from ux_copy import BAND_COPY, SCORE_COPY

TEMPLATE_DIR = Path(__file__).parent / "templates"

_env = Environment(
    loader=FileSystemLoader(str(TEMPLATE_DIR)),
    autoescape=select_autoescape(["html"]),
)


def render_html_report(audit_data: dict, graph_cap: int = 250) -> str:
    template = _env.get_template("report.html")
    return template.render(
        data=audit_data, data_json=json.dumps(audit_data), graph_cap=graph_cap,
        score_copy=SCORE_COPY, band_copy=BAND_COPY,
    )


def export_json(audit_data: dict) -> bytes:
    return json.dumps(audit_data, indent=2).encode("utf-8")


def export_csv(audit_data: dict) -> bytes:
    buf = io.StringIO()
    fieldnames = [
        "url", "status_code", "title", "word_count", "path_depth", "click_depth",
        "is_thin_content", "is_duplicate_of", "images_total", "images_missing_alt",
        "has_schema_org", "canonical", "internal_links_out_count", "script_count",
        "external_script_count", "error",
    ]
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in audit_data["pages"].values():
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")


def export_xlsx(audit_data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    HEADER_FILL = PatternFill(start_color="C1531F", end_color="C1531F", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    TITLE_FONT = Font(bold=True, size=14)
    WRAP = Alignment(wrap_text=True, vertical="top")
    SEV_FILLS = {
        0: PatternFill(start_color="E4EFE7", end_color="E4EFE7", fill_type="solid"),
        1: PatternFill(start_color="FBF0DA", end_color="FBF0DA", fill_type="solid"),
        2: PatternFill(start_color="F7E1B0", end_color="F7E1B0", fill_type="solid"),
        3: PatternFill(start_color="F3CFC9", end_color="F3CFC9", fill_type="solid"),
        4: PatternFill(start_color="B23A34", end_color="B23A34", fill_type="solid"),
    }
    PRIORITY_FILLS = {
        "high": PatternFill(start_color="F3CFC9", end_color="F3CFC9", fill_type="solid"),
        "medium": PatternFill(start_color="F7E1B0", end_color="F7E1B0", fill_type="solid"),
        "low": PatternFill(start_color="E4EFE7", end_color="E4EFE7", fill_type="solid"),
    }

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = f"A{row + 1}"

    def add_table(ws, name: str, n_rows: int, n_cols: int, header_row: int = 1):
        if n_rows < 1:
            return
        ref = f"A{header_row}:{get_column_letter(n_cols)}{header_row + n_rows}"
        table = Table(displayName=name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)

    def autosize(ws, widths: list[int]):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = Workbook()

    # ---- Overview ----
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "UX & IA Audit — Overview"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")
    meta_rows = [
        ("Site audited", audit_data["meta"]["start_url"]),
        ("Pages crawled", audit_data["meta"]["pages_crawled"]),
        ("Crawl finished", audit_data["meta"]["finished_at"][:19].replace("T", " ")),
        ("", ""),
        ("Overall UX Maturity", f"{audit_data['scoring']['ux_maturity_score']} / 100 ({audit_data['scoring']['ux_maturity_band']})"),
        ("Information Architecture", f"{audit_data['scoring']['ia_health_score']} / 100"),
        ("Content Quality", f"{audit_data['scoring']['content_quality_score']} / 100"),
        ("Accessibility", f"{audit_data['scoring']['accessibility_score']} / 100"),
        ("SEO / Findability", f"{audit_data['scoring']['seo_score']} / 100"),
        ("", ""),
        ("Orphan pages", audit_data["ia"]["orphan_page_count"]),
        ("Thin content pages", audit_data["content"]["thin_content_count"]),
        ("Duplicate content pages", audit_data["content"]["duplicate_content_page_count"]),
        ("Pages with accessibility issues", f"{audit_data['accessibility']['pages_with_issues']} / {audit_data['accessibility']['pages_analyzed']}"),
    ]
    for i, (label, val) in enumerate(meta_rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val)
    autosize(ws, [28, 46])

    # ---- Heuristic Evaluation ----
    ws2 = wb.create_sheet("Heuristic Evaluation")
    headers = ["Heuristic", "Assessed?", "Status", "Top Finding"]
    ws2.append(headers)
    for h in audit_data["heuristics"]:
        if not h["assessable"]:
            status, top_finding = "Not assessed", h["why_not"]
        elif h["findings"]:
            status = f"Severity {h['max_severity']} — {len(h['findings'])} finding(s)"
            top_finding = h["findings"][0]["text"]
        else:
            status, top_finding = "No issues found", ""
        row = [f"{h['id'].upper()} — {h['name']}", "Yes" if h["assessable"] else "No", status, top_finding]
        ws2.append(row)
        if h["assessable"] and h["findings"]:
            ws2.cell(row=ws2.max_row, column=3).fill = SEV_FILLS.get(h["max_severity"], SEV_FILLS[1])
    style_header(ws2)
    for row in ws2.iter_rows(min_row=2):
        row[3].alignment = WRAP
    autosize(ws2, [34, 12, 30, 60])

    # ---- Action Plan ----
    ws3 = wb.create_sheet("Action Plan")
    ws3.append(["Priority", "Impact", "Effort", "Area", "Action"])
    for item in audit_data["scoring"]["action_plan"]:
        ws3.append([item["priority"].capitalize(), item.get("impact", ""), item.get("effort", ""), item["area"], item["action"]])
        ws3.cell(row=ws3.max_row, column=1).fill = PRIORITY_FILLS.get(item["priority"], PatternFill())
    style_header(ws3)
    for row in ws3.iter_rows(min_row=2):
        row[4].alignment = WRAP
    autosize(ws3, [12, 11, 11, 16, 76])
    if ws3.max_row > 1:
        add_table(ws3, "ActionPlan", ws3.max_row - 1, 5)

    # ---- Keywords ----
    ws4 = wb.create_sheet("Keywords")
    ws4.append(["Keyword", "Occurrences", "Pages Found On", "% of Pages"])
    for kw in audit_data.get("keywords", {}).get("top_keywords", []):
        ws4.append([kw["term"], kw["count"], kw["pages"], kw["pct_of_pages"] / 100])
        ws4.cell(row=ws4.max_row, column=4).number_format = "0.0%"
    style_header(ws4)
    autosize(ws4, [26, 14, 16, 12])
    ws4["F1"] = "Top Phrases"
    ws4["F1"].font = Font(bold=True)
    ws4["G1"] = "Occurrences"
    ws4["G1"].font = Font(bold=True)
    for i, ph in enumerate(audit_data.get("keywords", {}).get("top_phrases", []), start=2):
        ws4.cell(row=i, column=6, value=ph["term"])
        ws4.cell(row=i, column=7, value=ph["count"])
    ws4.column_dimensions["F"].width = 26
    ws4.column_dimensions["G"].width = 14

    # ---- Integrations ----
    ws_int = wb.create_sheet("Integrations")
    ws_int.append(["Integration", "Category", "Pages Found On", "% of Pages"])
    for d in audit_data.get("integrations", {}).get("detected", []):
        ws_int.append([d["name"], d["category"], d["pages_found_on"], d["pct_of_pages"] / 100])
        ws_int.cell(row=ws_int.max_row, column=4).number_format = "0.0%"
    style_header(ws_int)
    autosize(ws_int, [28, 22, 16, 12])
    ws_int["F1"] = "Other scripts (unrecognized)"
    ws_int["F1"].font = Font(bold=True)
    ws_int["G1"] = "References"
    ws_int["G1"].font = Font(bold=True)
    for i, s in enumerate(audit_data.get("integrations", {}).get("other_scripts", []), start=2):
        ws_int.cell(row=i, column=6, value=s["domain"])
        ws_int.cell(row=i, column=7, value=s["reference_count"])
    ws_int.column_dimensions["F"].width = 34
    ws_int.column_dimensions["G"].width = 14

    # ---- Feature Matrix ----
    fm = audit_data.get("feature_matrix", {})
    if fm.get("rows"):
        ws_fm = wb.create_sheet("Feature Matrix")
        ws_fm.append(["Feature", "Detected?", "Pages Found On"])
        for row in fm["rows"]:
            ws_fm.append([row["name"], "Yes" if row["present"] else "No", row.get("page_count") or ""])
            if row["present"]:
                ws_fm.cell(row=ws_fm.max_row, column=2).fill = PatternFill(start_color="E4EFE7", end_color="E4EFE7", fill_type="solid")
        style_header(ws_fm)
        autosize(ws_fm, [32, 12, 16])

    # ---- Journey Map ----
    jm = audit_data.get("journey_map", {})
    if jm.get("journeys"):
        ws_jm = wb.create_sheet("Journey Map")
        ws_jm.append(["Persona", "Stage", "Found?", "Pages", "Closest Example", "Click Depth"])
        for journey in jm["journeys"]:
            for s in journey["stages"]:
                ws_jm.append([
                    journey["name"], s["name"], "Yes" if s["present"] else "No",
                    s.get("page_count") or "", s.get("example_url") or "",
                    s.get("click_depth") if s.get("click_depth") is not None else "",
                ])
        style_header(ws_jm)
        autosize(ws_jm, [24, 22, 10, 8, 46, 12])

    # ---- External Link Health ----
    link_health = audit_data.get("link_health", {})
    if link_health.get("checked"):
        ws_link = wb.create_sheet("External Link Health")
        ws_link.append(["Broken URL", "Status", "Linked From (count)", "Example Linking Page"])
        for b in link_health.get("broken", []):
            ws_link.append([b.get("url"), b.get("status_code") or b.get("error"), b.get("linked_from_count"), b.get("example_linking_page")])
        style_header(ws_link)
        autosize(ws_link, [46, 12, 18, 46])

    # ---- Page Inventory ----
    ws5 = wb.create_sheet("Page Inventory")
    fieldnames = [
        "url", "status_code", "title", "word_count", "path_depth", "click_depth",
        "is_thin_content", "is_duplicate_of", "images_total", "images_missing_alt",
        "has_schema_org", "canonical", "internal_links_out_count", "script_count",
        "external_script_count", "error",
    ]
    headers5 = ["URL", "Status", "Title", "Words", "Path Depth", "Click Depth", "Thin?",
                "Duplicate Of", "Images", "Missing Alt", "Has Schema?", "Canonical", "Internal Links Out",
                "Scripts", "External Scripts", "Error"]
    ws5.append(headers5)
    for row in audit_data["pages"].values():
        ws5.append([row.get(f) for f in fieldnames])
    style_header(ws5)
    autosize(ws5, [46, 9, 38, 8, 11, 11, 7, 30, 8, 11, 11, 30, 16, 9, 14, 16])
    if ws5.max_row > 1:
        add_table(ws5, "PageInventory", ws5.max_row - 1, len(headers5))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
